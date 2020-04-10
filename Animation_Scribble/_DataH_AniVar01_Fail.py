#-*-coding:utf-8-*-

import numpy as np
from openpyxl import *
import matplotlib.pyplot as plt
import matplotlib.animation as animation
import time

outs = 10
# outs denotes how many actual data is recorded as the output (10 means every 10th step is recorded

def write(values,varName,name,wCmd): #wCmd is a tuple of strings
    print ('Writing...')
    if 'csv' in wCmd:
        writeCSV(values,varName,name,s=outs)
    if 'xlsx' in wCmd:
        writeXLSX2(values,varName,name,s=outs)
    if 'plot' in wCmd:
        writePLOT(values,varName,name,s=outs)
    if 'animate' in wCmd:
        animateF(values,varName,name,s=outs)
    return values

def writeCSV(values,varName,name,s=1):
    S = 't'
    for i in varName:
        S += ',%s'%(i)
    
    for i in values[::s]:
        S+= '\n'+str(i[0])
        for j in i[1]:
            S += ','+str(j)
    f = open(name+'.csv','w')
    f.write(S)
    f.close()
    #print values
    print ('\nCSV Output Successful\nName >> %s'%(name+'.csv'))

def spl(values,varName,s=1):
    L = dict()
    L['t'] = []
    for i in varName:
        L[i] = []
    for i in values[::s]:
        L['t'].append(i[0])
        for j in range(len(i[1])):
            L[varName[j]].append(i[1][j])
    '''
    X = spl(values,varName)
    print (X.keys())
    for i in X.keys():
        print (X[i][:10])
    '''
    timestep = L['t'][1]-L['t'][0]
    return (L,timestep,)
    
def rect(values,varName,incN=True): #incN determines to include the name or not
    X = [tuple(['t']+list(varName))]
    for i in values:
        temp = []
        temp.append(i[0])
        for j in i[1]:
            temp.append(j)
        X.append(temp)
    return X
    
def rectXLSX(ws,rect):
    for i in range(len(rect)):
        for j in range(len(rect[i])):
            ws.cell(row=1+i,column=1+j).value=rect[i][j]
    return ws

def writeXLSX2(values,varName,name,s=1):
    wb = Workbook()
    ws = wb.active
    ws.title = name[:30] #worksheet title cannot be longer than 31 characters!
    rectXLSX(ws,rect(values,varName))
    wb.save(name+'.xlsx')
    print ('\nEXCEL Output Successful\nName >> %s'%(name+'.xlsx'))
    return wb

def writeXLSX(values,varName,name,s=1):
    #print (varName)
    print (rect(values,varName)[:10])
    wb = Workbook()
    ws = wb.active
    ws.title = name[:30] #worksheet title cannot be longer than 31 characters!
    #ws.font = fontStyle 
    ws.cell(row=1,column=1).value = 't'
    #ws.cell(row=1,column=1).font = fontStyle
    for i in range(len(varName)):
        ws.cell(row = 1,column = i+2).value = varName[i]
        #ws.cell(row = 1,column = i+2).font = fontStyle
    #print (values[:10])
    for i in range(len(values)):
        ws.cell(row=i+2,column=1).value = values[i][0]
        for j in range(len(values[i][1])):
            ws.cell(row=i+2,column=j+2).value = values[i][1][j]
    wb.save(name+'.xlsx')
    print ('\nEXCEL Output Successful\nName >> %s'%(name+'.xlsx'))    
    return wb

def writePLOT(values,varName,name,s=1,x = ['t'],y = ()):
    D = spl(values,varName,s)[0]
    if y == ():
        y = varName
    while len(y)>len(x):
        x += x[0]
    if len(x) == 1:
        plt.xlabel(x[0])
    for i in range(len(y)):
        plt.plot(D[x[i]],D[y[i]],label = '%s - %s' %(y[i],x[i]) )
    t = ''
    t += name+'\n'
    for i in y:
        t += i + ', '
    t = t[:-2] + ' - '
    for j in x:
        t += j + ', '
    plt.title(t[:-2] + ' graph')
    plt.grid(b=None, which='major', axis='both')
    plt.legend()
    plt.show()

def capture(vSpl,ts,te,timestep):
    iS, iE = int(ts/timestep), int(te/timestep)

def animateF(values,varName,name,s=1,x = 't',y = (),scale = 1, tail = 1): #scale governs the scale ratio between the real and animated time, and tail governs the 'time length' of tail
    L,timestep = spl(values,varName)
    fig = plt.figure()
    axes = plt.gca()
    line, = axes.plot([], [], 'r-')    

    if tail == 0:
        #print ('tail zero')
        def ani(t):
            line.set_xdata(L['t'][:int(t/timestep)])
            line.set_ydata(L[varName[0]][:int(t/timestep)])
            return line,
    else:
        def ani(t):
            x = t-tail
            if x<0:
                x = 0
            line.set_xdata(L['t'][int((x)/timestep):int(t/timestep)])
            line.set_ydata(L[varName[0]][int((x)/timestep):int(t/timestep)])
            return line,
    axes.set_autoscale_on(False)
    axes.set_xlim(min(L['t']),max(L['t']))
    axes.set_ylim(min(L[varName[0]]),max(L[varName[0]]))
    
    quality = 10
    anchor = np.arange(0,L['t'][-1],1/quality)
    intV = 0.01
    plt.show()
    for i in anchor:
        ani(i)
        plt.draw()
        plt.pause(1e-17)
        time.sleep(intV)
    
    time.sleep(2)
    plt.close(fig)

   
if __name__ == '__main__':
    pass
'''
###############################################################3
    ysample = random.sample(range(-50, 50), 100)
    
    xdata = []
    ydata = []
    plt.show()
    axes = plt.gca()
    axes.set_xlim(0, 100)
    axes.set_ylim(-50, +50)
    line, = axes.plot(xdata, ydata, 'r-')
    
    for i in range(100):
        xdata.append(i)
        ydata.append(ysample[i])
        line.set_xdata(xdata)
        line.set_ydata(ydata)
        plt.draw()
        plt.pause(1e-17)
        time.sleep(0.01)
'''
'''
def animateF(values,varName,name,s=1,x = 't',y = (),scale = 1, tail = 0): #scale governs the scale ratio between the real and animated time, and tail governs the 'time length' of tail
    L,timestep = spl(values,varName)
    for i in L.keys():
        print (L[i][:10])
    print (timestep)
    fig, ax = plt.subplots()
    line, = ax.plot([],[])
    if tail == 0:
        #print ('tail zero')
        def ani(t):
            print ('ani initiated')
            
            line.set_xdata(L['t'][:int(t/timestep)])
            line.set_ydata(L[varName[0]][:int(t/timestep)])
            return line,
    else:
        def ani(t):
            x = t-tail
            if x<0:
                x = 0
            line.set_xdata(L['t'][int((x)/timestep):int(t/timestep)])
            line.set_ydata(L[varName[0]][int((x)/timestep):int(t/timestep)])
            return line,
    ax.set_autoscale_on(False)
    ax.set_xlim(min(L['t']),max(L['t']))
    ax.set_ylim(min(L[varName[0]]),max(L[varName[0]]))
    
    quality = 10
    anchor = np.arange(0,L['t'][-1],1/quality)
    #print (np.arange(0,L['t'][-1],10/quality)[0])
    intV = 100
    #print ('hello')
    print (anchor,intV,ani)
    anim = animation.FuncAnimation(fig, ani, anchor,interval = intV)
    
    print (anim)
    plt.show()
    for i in anchor:
        ani(i)
        plt.pause(1e-17)
        time.sleep(intV)
'''