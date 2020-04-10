#-*-coding:utf-8-*-

import numpy as np
from openpyxl import *
import matplotlib.pyplot as plt
import matplotlib.animation as animation
import time
import copy

outs = 10
# outs denotes how many actual data is recorded as the output (10 means every 10th step is recorded


def write(values,varName,name='Untitled',wCmd=()): #wCmd is a tuple of strings
    print ('Initiate Output Module ...\n')
    if 'csv' in wCmd:
        writeCSV(values,varName,name,s=outs)
    if 'xlsx' in wCmd:
        writeXLSX2(values,varName,name,s=outs)
    if 'plot' in wCmd:
        writePLOT(values,varName,name,s=outs)
    if 'animate' in wCmd:
        animateF(values,varName,name,s=outs)
    return values

def writeCSV(values,varName,name,s=1):
    print ('Writing CSV file...\n')
    S = 't'
    for i in varName:
        S += ',%s'%(i)
    
    for i in values[::s]:
        S+= '\n'+str(i[0])
        for j in i[1]:
            S += ','+str(j)
    f = open(name+'.csv','w')
    f.write(S)
    f.close()
    #print values
    print ('\nCSV Output Successful\nName >> %s'%(name+'.csv'))

def spl(values,varName,s=1,incT=0): #incT denotes the new desired timestep
    if incT !=0:
        s = int(incT/(values[1][0]-values[0][0]))
    L = dict()
    L['t'] = []
    for i in varName:
        L[i] = []
    for i in values[::s]:
        L['t'].append(i[0])
        for j in range(len(i[1])):
            L[varName[j]].append(i[1][j])
    timestep = L['t'][1]-L['t'][0]
    return (L,timestep,)
    
def rect(values,varName,incN=True): #incN determines to include the name or not
    X = [tuple(['t']+list(varName))]
    for i in values:
        temp = []
        temp.append(i[0])
        for j in i[1]:
            temp.append(j)
        X.append(temp)
    return X
    
def rectXLSX(ws,rect):
    for i in range(len(rect)):
        for j in range(len(rect[i])):
            ws.cell(row=1+i,column=1+j).value=rect[i][j]
    return ws

def writeXLSX2(values,varName,name,s=1):
    print ('Writing XLSX file...\n')
    wb = Workbook()
    ws = wb.active
    ws.title = name[:30] #worksheet title cannot be longer than 31 characters!
    rectXLSX(ws,rect(values,varName))
    wb.save(name+'.xlsx')
    print ('\nEXCEL Output Successful\nName >> %s'%(name+'.xlsx'))
    return wb

def writeXLSX(values,varName,name,s=1):
    print ('Writing XLSX file...\n')
    #print (varName)
    #print (rect(values,varName)[:10])
    wb = Workbook()
    ws = wb.active
    ws.title = name[:30] #worksheet title cannot be longer than 31 characters!
    #ws.font = fontStyle 
    ws.cell(row=1,column=1).value = 't'
    #ws.cell(row=1,column=1).font = fontStyle
    for i in range(len(varName)):
        ws.cell(row = 1,column = i+2).value = varName[i]
        #ws.cell(row = 1,column = i+2).font = fontStyle
    #print (values[:10])
    for i in range(len(values)):
        ws.cell(row=i+2,column=1).value = values[i][0]
        for j in range(len(values[i][1])):
            ws.cell(row=i+2,column=j+2).value = values[i][1][j]
    wb.save(name+'.xlsx')
    print ('\nEXCEL Output Successful\nName >> %s'%(name+'.xlsx'))    
    return wb

def writePLOT(values,varName,name,s=1,x = ['t'],y = ()):
    print ('Plotting Result...\n')    
    D = spl(values,varName,s)[0]
    if y == ():
        y = varName
    while len(y)>len(x):
        x += x[0]
    if len(x) == 1:
        plt.xlabel(x[0])
    for i in range(len(y)):
        plt.plot(D[x[i]],D[y[i]],label = '%s - %s' %(y[i],x[i]) )
    t = ''
    t += name+'\n'
    for i in y:
        t += i + ', '
    t = t[:-2] + ' - '
    for j in x:
        t += j + ', '
    plt.title(t[:-2] + ' graph')
    plt.grid(b=None, which='major', axis='both')
    plt.legend()
    plt.show()
    print ('Plotting Complete')

def animateF(values,varName,name,s=1,subj = [],scale = 1, tail = 1,conS = ()): 
    print ('Prepairing Animation...\n')

    #scale governs the scale ratio between the real and animated time, and tail governs the 'time length' of tail
    #subj governs the things to plot ex) (('x','y'),('t','x')) versus time
    #conn connects the respective places. (putting 'O' includes the origin as the starting point)
    
    #default subject graphs everything agianst time with respect to time
    if subj == []:
        for i in varName:
            subj.append(('t',i))
    
    # initialize object necessary
    L,timestep = spl(values,varName,incT = 0.01*scale)
    #print (timestep)
    #initialize matplotlib objects
    fig = plt.figure()
    axes = plt.gca()
    
    # initialize line variables
    Lines = []
    con = False
    coz = False
    if conS != ():
        con = True
        if '0' in conS:
            coz = True
        link, = axes.plot([],[],label = 'link')
    for i in subj:
        Lines.append(axes.plot([], [], label = '(%s,%s)vs Time'%(i[0],i[1])))
#    line, = Lines[0]
#    print (type(line))
    
    if tail == 0:
        #print ('tail zero')
        def ani(t):
            for i in range(len(Lines)):
                line, = Lines[i]
                line.set_xdata(L[subj[i][0]][:int(t/timestep)])
                line.set_ydata(L[subj[i][1]][:int(t/timestep)])
            # Set up the links
            if con:
                x,y = [],[]
                if coz:
                    x,y = [0],[0]
                for i in conS:
                    if type(i) == int:
                        x.append(L[subj[i][0]][int(t/timestep)])
                        y.append(L[subj[i][1]][int(t/timestep)])
                link.set_xdata(x)
                link.set_ydata(y)

    else:
        def ani(t):
            x = t-tail
            if x<0:
                x = 0
            for i in range(len(Lines)):
                line, = Lines[i]
                line.set_xdata(L[subj[i][0]][int(x/timestep):int(t/timestep)])
                line.set_ydata(L[subj[i][1]][int(x/timestep):int(t/timestep)])
            #Set up the links
            if con:
                x,y = [],[]
                if coz:
                    x,y = [0],[0]
                for i in conS:
                    if type(i) == int:
                        x.append(L[subj[i][0]][int(t/timestep)])
                        y.append(L[subj[i][1]][int(t/timestep)])
                link.set_xdata(x)
                link.set_ydata(y)
    
    # determines proper x, y lim
    axes.set_autoscale_on(False)
    X = []
    Y = []
    #for i in L.keys():
    #    print (i,L[i][:10])
    #print (subj)
    for i in subj:
        X += L[i[0]]
        Y += L[i[1]]
    minX,minY,maxX,maxY = min(X),min(Y),max(X),max(Y)
    if subj[0][0] == 't':
        axes.set_xlim(min(L['t']),max(L['t']))
        axes.set_ylim(minY-abs(minY*0.05),maxY+abs(maxY*0.05))
    else:
        axes.set_xlim(minX-abs(minX*0.05),maxX+abs(maxX*0.05))
        axes.set_ylim(minY-abs(minY*0.05),maxY+abs(maxY*0.05))

    # the initialization and formats are in place. now for the animation.
    quality = 10/scale
    anchor = np.arange(0,L['t'][-1],1/quality)
    intV = 0.01
    plt.show()
    plt.title(name+'\nAnimation %s vs time'%(str(subj)))
    plt.legend()
    try:
        for i in anchor:
            ani(i)
            plt.draw()
            plt.pause(1e-17)
            time.sleep(intV)
    except:
        plt.close(fig)
        print ('Animation Closed \n')
        return
    time.sleep(2)
    plt.close(fig)
    print ('Animation Finished\n')
    

   
if __name__ == '__main__':
    pass